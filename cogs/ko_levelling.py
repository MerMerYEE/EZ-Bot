import discord
from  discord import File
from discord.ext import commands
import time
import wget
import sys
import re
import pandas as pd
import os
import string
from selenium import webdriver
import random
import urllib
import json
import urllib.request
from openpyxl import Workbook
import bs4
from PIL import Image, ImageDraw, ImageFont
import numpy as np
from urllib.request import urlopen, Request
import openpyxl
from requests import get

if os.path.isdir("./lib"):
    print("Lib Exist")
else:
    os.mkdir("./lib")
    print("Make Lib Folder")

if os.path.isdir("./lib/users"):
    print("Users Folder Exist")
else:
    os.mkdir("./lib/users")
    print("Make Users Folder")

class ko_Levelling(commands.Cog):

    def __init__(self, client):
        self.client = client

    # Events
    @commands.Cog.listener()
    async def on_message(self, ctx):
        #폴더 만들기
        if os.path.isdir("./lib/users"):
            print("Users Folder Exist")
        else:
            os.mkdir("./lib/users")
            print("Make Users Folder")

        #유저정보 생성
        if os.path.isfile("./lib/users/" + str(ctx.author.id) + ".xlsx"):
            print("User-File Exist")
        else:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Exp"
            ws.cell(row=2, column = 1).value = "Lvl"
            ws.cell(row=3, column = 1).value = "levelmax"
            ws.cell(row=4, column = 1).value = "currentexp"
            for i in range (1,5):
                ws.cell(row=i, column=2).value = "0"
            wb.save("./lib/users/" + str(ctx.author.id) + ".xlsx")
            wb.close()
            print(str(ctx.author.id) + "mk")

        #레벨 정리
        if os.path.isfile("./lib/users/" + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook("./lib/users/" + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            currentexp = float(ws.cell(row=1, column=2).value)
            crlvl = int(ws.cell(row=2, column=2).value)
            addvalue = random.randrange(1,8)
            currentexp = currentexp + (addvalue/8)
            nowexp = float(ws.cell(row=4, column=2).value)
            nowexp = nowexp + (addvalue/8)
            ws.cell(row=1, column=2).value = str(currentexp)
            expmax = crlvl**2 - 2*crlvl + 7
            ws.cell(row=3, column=2).value = str(expmax)
            if (nowexp >= expmax):
                crlvl = crlvl + 1
                nowexp = 0
            ws.cell(row=4, column=2).value = str(nowexp)
            ws.cell(row=2, column=2).value = str(crlvl)
            wb.save("./lib/users/" + str(ctx.author.id) + ".xlsx")
            wb.close()

    # Commands
    @commands.command()
    async def 레벨(self, ctx):
        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile("./lib/users/" + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook("./lib/users/" + str(user.id) + ".xlsx")
                    ws = wb.active
                    exp = ws.cell(row=4, column=2).value
                    lvl = ws.cell(row = 2, column = 2).value
                    expmax = ws.cell(row=3, column=2).value
                    base = Image.open("./src/base.png")
                    bar = Image.open("./src/bar.png")
                    mas = round(900*(float(exp)/int(expmax))-900)
                    base.paste(bar, (mas,0))
                    im = Image.open("./src/bg.png")
                    base.paste(im, (0,0), im)
                    font_folder = "./Font/"
                    with open("./src/avatar.png", "wb") as file:
                        response = get(user.avatar_url)
                        file.write(response.content)
                    im2 = Image.open("./src/avatar.png")
                    resize_image = im2.resize((256,256))
                    resize_image.save("./src/avatar2.png")
                    im3 = Image.open("./src/avatar2.png")
                    base.paste(im3, (50, 140))
                    font = ImageFont.truetype(os.path.join(font_folder, 'NotoSansCJKkr-DemiLight.otf'), 40)
                    font2 = ImageFont.truetype(os.path.join(font_folder, 'Uni Sans Heavy.otf'), 40)
                    draw = ImageDraw.Draw(base)
                    draw.text((500, 120), user.name, fill=(36, 36, 36, 255), font=font)
                    draw.text((500, 258), lvl, fill=(36, 36, 36, 255), font=font2)
                    draw.text((475, 377), exp + "/" + expmax, fill=(36, 36, 36, 255), font=font2)
                    base.save("./src/level.png")
                    await ctx.send(file = discord.File("./src/level.png"))
                else:
                    embed = discord.Embed(title="**에러!**", description= "데이터가 존재하지 않아요!", color=0x8680df)
                    embed.set_thumbnail(url="https://raw.githubusercontent.com/Shio7/EZ-Bot/master/images/shio_error.png")
                    embed.set_footer(text="Offered by NACL - Shio", icon_url="https://raw.githubusercontent.com/Shio7/EZ-Bot/master/images/Shio8.png")
                    await ctx.trigger_typing()
                    await ctx.send_file(embed=embed)

        else:
            if os.path.isfile("./lib/servers/" + str(ctx.guild.id) + "/" + str(ctx.author.id) + ".txt"):
                wb = openpyxl.load_workbook("./lib/users/" + str(ctx.author.id) + ".xlsx")
                ws = wb.active
                exp = ws.cell(row=4, column=2).value
                lvl = ws.cell(row = 2, column = 2).value
                expmax = ws.cell(row=3, column=2).value
                base = Image.open("./src/base.png")
                bar = Image.open("./src/bar.png")
                mas = round(900*(float(exp)/int(expmax))-900)
                base.paste(bar, (mas,0))
                im = Image.open("./src/bg.png")
                base.paste(im, (0,0), im)
                font_folder = "./Font/"
                with open("./src/avatar.png", "wb") as file:
                    response = get(ctx.author.avatar_url)
                    file.write(response.content)
                im2 = Image.open("./src/avatar.png")
                resize_image = im2.resize((256,256))
                resize_image.save("./src/avatar2.png")
                im3 = Image.open("./src/avatar2.png")
                base.paste(im3, (50, 140))
                font = ImageFont.truetype(os.path.join(font_folder, 'NotoSansCJKkr-DemiLight.otf'), 40)
                font2 = ImageFont.truetype(os.path.join(font_folder, 'Uni Sans Heavy.otf'), 40)
                draw = ImageDraw.Draw(base)
                draw.text((500, 120), ctx.author.name, fill=(36, 36, 36, 255), font=font)
                draw.text((500, 258), lvl, fill=(36, 36, 36, 255), font=font2)
                draw.text((475, 377), exp + "/" + expmax, fill=(36, 36, 36, 255), font=font2)
                base.save("./src/level.png")
                await ctx.send(file = discord.File("./src/level.png"))

            else:
                embed = discord.Embed(title="**에러!**", description= "데이터가 존재하지 않아요!", color=0x8680df)
                embed.set_thumbnail(url="https://raw.githubusercontent.com/Shio7/EZ-Bot/master/images/shio_error.png")
                embed.set_footer(text="Offered by NACL - Shio", icon_url="https://raw.githubusercontent.com/Shio7/EZ-Bot/master/images/Shio8.png")
                await ctx.trigger_typing()
                await ctx.send(embed=embed)




def setup(client):
    client.add_cog(ko_Levelling(client))
